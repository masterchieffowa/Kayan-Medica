# system.py
import csv
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from datetime import datetime
from collections import defaultdict
import os
from PIL import Image, ImageTk  # requires pillow

FILE_PATH = "db.xlsx"
LOGO_PATH = "logo.png"   # Put your logo here (same folder); optional

# ------------------- Ensure workbook and sheets -------------------
def ensure_workbook():
    if not os.path.exists(FILE_PATH):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        inv = wb.create_sheet("Inventory (المخزون)")
        inv.append(["رقم", "اسم الصنف", "الكمية الحالية", "الكمية الكاملة",
                    "المستورد", "سعر الشراء", "سعر البيع", "اسم المخزن", "تاريخ الإدخال", "ملاحظات"])
        sales = wb.create_sheet("Sales (المبيعات)")
        sales.append(["رقم الفاتورة", "التاريخ", "الصنف", "الكمية", "المشتري",
                      "السعر الكلي", "صافي الربح", "طريقة الدفع", "ملاحظات"])
        pur = wb.create_sheet("Purchases (المشتريات)")
        pur.append(["الرقم", "الصنف", "الكمية المضافة", "المستورد",
                    "تاريخ الشراء", "التكلفة الإجمالية", "سعر الشراء", "هامش_فردي", "اسم المخزن", "ملاحظات"])
        wb.save(FILE_PATH)

ensure_workbook()
wb = openpyxl.load_workbook(FILE_PATH)
inventory_sheet = wb["Inventory (المخزون)"]
sales_sheet = wb["Sales (المبيعات)"]
purchases_sheet = wb["Purchases (المشتريات)"]

def save_excel():
    wb.save(FILE_PATH)

# ------------------- helpers -------------------
def safe_int(v, default=0):
    try:
        return int(float(v))
    except Exception:
        return default

def safe_float(v, default=0.0):
    try:
        return float(v)
    except Exception:
        return default

def load_products():
    products = {}
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        if row[1]:
            name = str(row[1])
            products[name] = {
                "id": row[0],
                "name": name,
                "qty": safe_int(row[2]),
                "full_qty": safe_int(row[3]),
                "importer": row[4],
                "purchase": safe_float(row[5]),
                "sell": safe_float(row[6]),
                "store": row[7] if len(row) > 7 else "",
                "date": row[8] if len(row) > 8 else "",
                "notes": row[9] if len(row) > 9 else ""
            }
    return products

def refresh_combobox_values(combo):
    combo['values'] = list(load_products().keys())

def insert_inventory_row(item):
    new_id = inventory_sheet.max_row
    now = datetime.now().strftime("%Y-%m-%d")
    inventory_sheet.append([new_id, item['name'], item['qty'], item.get('full_qty', item['qty']),
                            item.get('importer', ""), item.get('purchase', 0.0), item.get('sell', 0.0),
                            item.get('store', ""), now, item.get('notes', "")])
    save_excel()

# ------------------- GUI -------------------
root = tk.Tk()
root.title("نظام إدارة المبيعات والمخزون - متقدّم")
root.geometry("1100x700")

# main frames: left = notebook, right = image/info
main_frame = tk.Frame(root)
main_frame.pack(fill="both", expand=True)

left_frame = tk.Frame(main_frame)
left_frame.pack(side="left", fill="both", expand=True)

right_frame = tk.Frame(main_frame, width=300)
right_frame.pack(side="right", fill="y")

# load logo if exists
logo_label = tk.Label(right_frame)
logo_label.pack(padx=10, pady=10)
if os.path.exists(LOGO_PATH):
    try:
        pil_img = Image.open(LOGO_PATH)
        pil_img.thumbnail((260, 260))
        tk_img = ImageTk.PhotoImage(pil_img)
        logo_label.configure(image=tk_img)
        logo_label.image = tk_img
    except Exception as e:
        logo_label.configure(text="Logo load error")
else:
    logo_label.configure(text="ضع الشعار هنا (logo.png)")

# small info area under image
info_text = tk.Text(right_frame, width=34, height=20)
info_text.pack(padx=8, pady=6)
info_text.insert("1.0", "معلومات البرنامج\n\n- يمكنك إضافة/تحديث مخزون.\n- تسجيل مشتريات ومبيعات.\n- التقارير بتدعم نطاق زمني.\n")
info_text.config(state="disabled")

notebook = ttk.Notebook(left_frame)
notebook.pack(fill="both", expand=True)

# ------------------- Inventory Tab -------------------
frame_inv = tk.Frame(notebook)
notebook.add(frame_inv, text="المخزون (Inventory)")

inv_controls = tk.Frame(frame_inv)
inv_controls.pack(fill="x", padx=8, pady=6)

tk.Label(inv_controls, text="اسم الصنف:").grid(row=0, column=0, sticky="e")
inv_name = tk.StringVar()
tk.Entry(inv_controls, textvariable=inv_name, width=30).grid(row=0, column=1)

tk.Label(inv_controls, text="الكمية:").grid(row=0, column=2, sticky="e")
inv_qty = tk.StringVar()
tk.Entry(inv_controls, textvariable=inv_qty, width=12).grid(row=0, column=3)

tk.Label(inv_controls, text="سعر الشراء:").grid(row=1, column=0, sticky="e")
inv_purchase = tk.StringVar()
tk.Entry(inv_controls, textvariable=inv_purchase, width=12).grid(row=1, column=1)

tk.Label(inv_controls, text="سعر البيع:").grid(row=1, column=2, sticky="e")
inv_sell = tk.StringVar()
tk.Entry(inv_controls, textvariable=inv_sell, width=12).grid(row=1, column=3)

tk.Label(inv_controls, text="المستورد:").grid(row=2, column=0, sticky="e")
inv_importer = tk.StringVar()
tk.Entry(inv_controls, textvariable=inv_importer, width=30).grid(row=2, column=1)

tk.Label(inv_controls, text="اسم المخزن:").grid(row=2, column=2, sticky="e")
inv_store = tk.StringVar()
tk.Entry(inv_controls, textvariable=inv_store, width=12).grid(row=2, column=3)

tk.Label(inv_controls, text="ملاحظات:").grid(row=3, column=0, sticky="e")
inv_notes = tk.StringVar()
tk.Entry(inv_controls, textvariable=inv_notes, width=60).grid(row=3, column=1, columnspan=3, sticky="w")

inv_editing_item = {}

def inv_clear_inputs():
    inv_name.set("")
    inv_qty.set("")
    inv_purchase.set("")
    inv_sell.set("")
    inv_importer.set("")
    inv_store.set("")
    inv_notes.set("")
    inv_editing_item.clear()
    inv_add_btn.config(text="إضافة صنف")

def refresh_inventory_tree():
    for row in inv_tree.get_children():
        inv_tree.delete(row)
    for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
        if row[1]:
            inv_tree.insert("", "end", values=row)

def inv_add_or_update():
    name = inv_name.get().strip()
    if not name:
        messagebox.showerror("خطأ", "أدخل اسم الصنف")
        return
    qty = safe_int(inv_qty.get(), 0)
    purchase = safe_float(inv_purchase.get(), 0.0)
    sell = safe_float(inv_sell.get(), 0.0)
    importer = inv_importer.get().strip()
    store = inv_store.get().strip()
    notes = inv_notes.get().strip()
    if inv_editing_item.get("name"):
        target_name = inv_editing_item["name"]
        for row in inventory_sheet.iter_rows(min_row=2):
            if str(row[1].value) == target_name:
                row[1].value = name
                row[2].value = qty
                if not row[3].value:
                    row[3].value = qty
                row[4].value = importer
                row[5].value = purchase
                row[6].value = sell
                row[7].value = store
                row[9].value = notes
                break
        save_excel()
        messagebox.showinfo("تم", "تم تحديث الصنف")
    else:
        item = {"name": name, "qty": qty, "purchase": purchase, "sell": sell,
                "importer": importer, "store": store, "notes": notes}
        insert_inventory_row(item)
        messagebox.showinfo("تم", "تم إضافة صنف جديد")
    inv_clear_inputs()
    refresh_inventory_tree()
    refresh_combobox_values(sales_combo)
    refresh_purchases_table()

tk.Button(inv_controls, text="إضافة/تحديث", command=inv_add_or_update, bg="#2c7", width=15).grid(row=4, column=1, pady=8)
tk.Button(inv_controls, text="مسح الحقول", command=inv_clear_inputs, width=15).grid(row=4, column=2, pady=8)

inv_cols = ["رقم", "اسم الصنف", "الكمية الحالية", "الكمية الكاملة", "المستورد", "سعر الشراء", "سعر البيع", "اسم المخزن", "تاريخ الإدخال", "ملاحظات"]
inv_tree = ttk.Treeview(frame_inv, columns=inv_cols, show="headings", height=12)
for c in inv_cols:
    inv_tree.heading(c, text=c)
    inv_tree.column(c, width=100, anchor="center")
inv_tree.pack(fill="both", expand=True, padx=8, pady=6)
refresh_inventory_tree()

def on_inv_select(event):
    sel = inv_tree.focus()
    if not sel:
        return
    vals = inv_tree.item(sel, "values")
    inv_editing_item.clear()
    inv_editing_item["name"] = str(vals[1])
    inv_name.set(vals[1])
    inv_qty.set(vals[2])
    inv_purchase.set(vals[5])
    inv_sell.set(vals[6])
    inv_importer.set(vals[4])
    inv_store.set(vals[7])
    inv_notes.set(vals[9] if len(vals) > 9 else "")

inv_tree.bind("<<TreeviewSelect>>", on_inv_select)

# ------------------- Purchases Tab -------------------
frame_pur = tk.Frame(notebook)
notebook.add(frame_pur, text="المشتريات (Purchases)")

pur_inputs = tk.Frame(frame_pur)
pur_inputs.pack(fill="x", padx=8, pady=6)

tk.Label(pur_inputs, text="اسم الصنف:").grid(row=0, column=0, sticky="e")
pur_name = tk.StringVar()
tk.Entry(pur_inputs, textvariable=pur_name, width=30).grid(row=0, column=1)

tk.Label(pur_inputs, text="الكمية المضافة:").grid(row=0, column=2, sticky="e")
pur_qty = tk.StringVar()
tk.Entry(pur_inputs, textvariable=pur_qty, width=12).grid(row=0, column=3)

tk.Label(pur_inputs, text="سعر الشراء للوحدة:").grid(row=1, column=0, sticky="e")
pur_price = tk.StringVar()
tk.Entry(pur_inputs, textvariable=pur_price, width=12).grid(row=1, column=1)

tk.Label(pur_inputs, text="هامش % (اختياري):").grid(row=1, column=2, sticky="e")
pur_margin = tk.StringVar()
tk.Entry(pur_inputs, textvariable=pur_margin, width=12).grid(row=1, column=3)

tk.Label(pur_inputs, text="المستورد:").grid(row=2, column=0, sticky="e")
pur_importer = tk.StringVar()
tk.Entry(pur_inputs, textvariable=pur_importer, width=20).grid(row=2, column=1)

tk.Label(pur_inputs, text="اسم المخزن:").grid(row=2, column=2, sticky="e")
pur_store = tk.StringVar()
tk.Entry(pur_inputs, textvariable=pur_store, width=20).grid(row=2, column=3)

tk.Label(pur_inputs, text="ملاحظات:").grid(row=3, column=0, sticky="e")
pur_notes = tk.StringVar()
tk.Entry(pur_inputs, textvariable=pur_notes, width=60).grid(row=3, column=1, columnspan=3, sticky="w")

pur_cols = ["الرقم", "الصنف", "الكمية المضافة", "المستورد", "تاريخ الشراء", "التكلفة الإجمالية", "سعر الشراء", "هامش_فردي", "اسم المخزن", "ملاحظات"]
pur_tree = ttk.Treeview(frame_pur, columns=pur_cols, show="headings", height=12)
for c in pur_cols:
    pur_tree.heading(c, text=c)
    pur_tree.column(c, width=110, anchor="center")
pur_tree.pack(fill="both", expand=True, padx=8, pady=6)

def refresh_purchases_table():
    for r in pur_tree.get_children():
        pur_tree.delete(r)
    for row in purchases_sheet.iter_rows(min_row=2, values_only=True):
        pur_tree.insert("", "end", values=row)

refresh_purchases_table()

def add_purchase():
    name = pur_name.get().strip()
    qty = safe_int(pur_qty.get(), 0)
    price = safe_float(pur_price.get(), 0.0)
    margin_pct = None
    if pur_margin.get().strip() != "":
        margin_pct = safe_float(pur_margin.get(), None)
    importer = pur_importer.get().strip()
    store = pur_store.get().strip()
    notes = pur_notes.get().strip()
    if not name or qty <= 0 or price <= 0:
        messagebox.showerror("خطأ", "تأكد من اسم الصنف والكمية وسعر الشراء")
        return
    products = load_products()
    now = datetime.now().strftime("%Y-%m-%d")
    total_cost = qty * price

    if name in products:
        # update existing
        for row in inventory_sheet.iter_rows(min_row=2):
            if str(row[1].value) == name:
                cur = safe_int(row[2].value, 0)
                row[2].value = cur + qty
                # update full qty
                full = safe_int(row[3].value, 0)
                row[3].value = full + qty if full else cur + qty
                row[4].value = importer or row[4].value
                row[5].value = price  # update purchase price to new purchase price
                # compute sell price using margin if provided; otherwise keep existing sell
                if margin_pct is not None:
                    sell_price = price * (1 + margin_pct / 100.0)
                    row[6].value = sell_price
                else:
                    # if sell blank, auto compute markup 30%
                    if not safe_float(row[6], 0.0):
                        row[6].value = price * 1.3
                row[7].value = store or row[7].value
                break
        # compute single margin per unit if sell available
        sell_price_now = load_products()[name]["sell"]
        single_margin = sell_price_now - price
    else:
        # create new inventory entry; if margin provided, compute sell accordingly
        if margin_pct is None:
            sell_price = price * 1.3
        else:
            sell_price = price * (1 + margin_pct / 100.0)
        new_item = {"name": name, "qty": qty, "purchase": price, "sell": sell_price,
                    "importer": importer, "store": store, "notes": notes}
        insert_inventory_row(new_item)
        single_margin = sell_price - price

    pur_id = purchases_sheet.max_row
    purchases_sheet.append([pur_id, name, qty, importer, now, total_cost, price, single_margin, store, notes])
    save_excel()

    # clear inputs
    pur_name.set("")
    pur_qty.set("")
    pur_price.set("")
    pur_margin.set("")
    pur_importer.set("")
    pur_store.set("")
    pur_notes.set("")

    refresh_purchases_table()
    refresh_inventory_tree()
    refresh_combobox_values(sales_combo)
    messagebox.showinfo("تم", f"تمت إضافة المشتريات. التكلفة: {total_cost} ، هامش فردي تقريبي: {single_margin}")

tk.Button(pur_inputs, text="تسجيل المشتريات", command=add_purchase, bg="#2b8cff", fg="white", width=20).grid(row=4, column=1, pady=8)

# ------------------- Sales Tab -------------------
frame_sales = tk.Frame(notebook)
notebook.add(frame_sales, text="المبيعات (Sales)")

sales_top = tk.Frame(frame_sales)
sales_top.pack(fill="x", padx=8, pady=6)

tk.Label(sales_top, text="اختر الصنف:").grid(row=0, column=0, sticky="e")
sales_product = tk.StringVar()
sales_combo = ttk.Combobox(sales_top, textvariable=sales_product, values=list(load_products().keys()), state="readonly", width=45)
sales_combo.grid(row=0, column=1, sticky="w")

tk.Label(sales_top, text="الكمية:").grid(row=0, column=2, sticky="e")
sales_qty = tk.StringVar()
tk.Entry(sales_top, textvariable=sales_qty, width=12).grid(row=0, column=3)

tk.Label(sales_top, text="المشتري:").grid(row=1, column=0, sticky="e")
sales_buyer = tk.StringVar()
tk.Entry(sales_top, textvariable=sales_buyer, width=30).grid(row=1, column=1, sticky="w")

tk.Label(sales_top, text="طريقة الدفع:").grid(row=1, column=2, sticky="e")
sales_pay = tk.StringVar()
tk.Entry(sales_top, textvariable=sales_pay, width=12).grid(row=1, column=3)

tk.Label(sales_top, text="سعر البيع (الوحدة):").grid(row=2, column=0, sticky="e")
sales_sell_display = tk.StringVar()
tk.Entry(sales_top, textvariable=sales_sell_display, state="readonly", width=18).grid(row=2, column=1, sticky="w")

tk.Label(sales_top, text="سعر الشراء (الوحدة):").grid(row=2, column=2, sticky="e")
sales_purchase_display = tk.StringVar()
tk.Entry(sales_top, textvariable=sales_purchase_display, state="readonly", width=18).grid(row=2, column=3, sticky="w")

def on_sales_product_select(event=None):
    name = sales_product.get()
    products = load_products()
    if name in products:
        sales_sell_display.set(products[name]["sell"])
        sales_purchase_display.set(products[name]["purchase"])
    else:
        sales_sell_display.set("")
        sales_purchase_display.set("")

sales_combo.bind("<<ComboboxSelected>>", on_sales_product_select)

sales_cols = ["رقم الفاتورة", "التاريخ", "الصنف", "الكمية", "المشتري", "السعر الكلي", "صافي الربح", "طريقة الدفع", "ملاحظات"]
sales_tree = ttk.Treeview(frame_sales, columns=sales_cols, show="headings", height=12)
for c in sales_cols:
    sales_tree.heading(c, text=c)
    sales_tree.column(c, width=110, anchor="center")
sales_tree.pack(fill="both", expand=True, padx=8, pady=6)

def refresh_sales_table():
    for r in sales_tree.get_children():
        sales_tree.delete(r)
    for row in sales_sheet.iter_rows(min_row=2, values_only=True):
        sales_tree.insert("", "end", values=row)

refresh_sales_table()

def add_sale():
    name = sales_product.get().strip()
    qty = safe_int(sales_qty.get(), 0)
    buyer = sales_buyer.get().strip()
    pay = sales_pay.get().strip() or "غير محدد"
    if not name or qty <= 0:
        messagebox.showerror("خطأ", "اختر الصنف وأدخل كمية صحيحة")
        return
    products = load_products()
    if name not in products:
        messagebox.showerror("خطأ", "الصنف غير موجود في المخزون")
        return
    prod = products[name]
    if qty > prod["qty"]:
        messagebox.showerror("خطأ", f"الكمية المطلوبة أكبر من المتاحة ({prod['qty']})")
        return
    sell_price = prod["sell"]
    purchase_price = prod["purchase"]
    total = qty * sell_price
    profit = (sell_price - purchase_price) * qty

    # update inventory
    for row in inventory_sheet.iter_rows(min_row=2):
        if str(row[1].value) == name:
            cur = safe_int(row[2].value, 0)
            row[2].value = cur - qty
            break

    sale_id = sales_sheet.max_row
    sales_sheet.append([sale_id, datetime.now().strftime("%Y-%m-%d"), name, qty, buyer, total, profit, pay, ""])
    save_excel()

    # clear
    sales_product.set("")
    sales_qty.set("")
    sales_buyer.set("")
    sales_pay.set("")
    sales_sell_display.set("")
    sales_purchase_display.set("")

    refresh_inventory_tree()
    refresh_sales_table()
    refresh_combobox_values(sales_combo)
    messagebox.showinfo("تم", f"سجل بيع بنجاح. السعر الكلي: {total} - صافي الربح: {profit}")

tk.Button(frame_sales, text="تسجيل بيع", command=add_sale, bg="#28a745", fg="white").pack(pady=6)

# ------------------- Reports Tab -------------------
frame_reports = tk.Frame(notebook)
notebook.add(frame_reports, text="التقارير (Reports)")

report_controls = tk.Frame(frame_reports)
report_controls.pack(fill="x", padx=8, pady=6)

tk.Label(report_controls, text="من تاريخ (YYYY-MM-DD):").grid(row=0, column=0, sticky="e")
report_from = tk.StringVar()
tk.Entry(report_controls, textvariable=report_from, width=15).grid(row=0, column=1)

tk.Label(report_controls, text="إلى تاريخ (YYYY-MM-DD):").grid(row=0, column=2, sticky="e")
report_to = tk.StringVar()
tk.Entry(report_controls, textvariable=report_to, width=15).grid(row=0, column=3)

tk.Label(report_controls, text="اسم الصنف (اختياري):").grid(row=1, column=0, sticky="e")
report_item = tk.StringVar()
tk.Entry(report_controls, textvariable=report_item, width=30).grid(row=1, column=1, columnspan=2, sticky="w")

report_table = ttk.Treeview(frame_reports, columns=("type", "item", "qty", "value", "date"), show="headings", height=20)
for col, label in zip(("type", "item", "qty", "value", "date"), ("النوع", "الصنف", "الكمية", "القيمة", "التاريخ")):
    report_table.heading(col, text=label)
    report_table.column(col, width=120, anchor="center")
report_table.pack(fill="both", expand=True, padx=8, pady=8)

def parse_date(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return None

def export_report_csv():
    file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
    if not file_path:
        return

    try:
        with open(file_path, mode='w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(["النوع", "الصنف", "الكمية", "القيمة", "التاريخ"])

            for row_id in report_table.get_children():
                row = report_table.item(row_id)['values']
                writer.writerow(row)

        messagebox.showinfo("تم", f"تم حفظ التقرير كـ CSV في:\n{file_path}")
    except Exception as e:
        messagebox.showerror("خطأ", f"حدث خطأ أثناء الحفظ:\n{e}")

def export_report_excel():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "التقرير"

        ws.append(["النوع", "الصنف", "الكمية", "القيمة", "التاريخ"])

        for row_id in report_table.get_children():
            row = report_table.item(row_id)['values']
            ws.append(row)

        wb.save(file_path)
        messagebox.showinfo("تم", f"تم حفظ التقرير كـ Excel في:\n{file_path}")
    except Exception as e:
        messagebox.showerror("خطأ", f"حدث خطأ أثناء الحفظ:\n{e}")

def refresh_report_table():
    report_table.delete(*report_table.get_children())
    from_date = parse_date(report_from.get())
    to_date = parse_date(report_to.get())
    item_filter = report_item.get().strip()

    # Purchases
    for row in purchases_sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[1])
        qty = safe_int(row[2])
        value = safe_float(row[5])  # total cost
        date_str = str(row[4])
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        except:
            continue
        if from_date and date_obj < from_date:
            continue
        if to_date and date_obj > to_date:
            continue
        if item_filter and item_filter not in name:
            continue
        report_table.insert("", "end", values=("شراء", name, qty, value, date_str))

    # Sales
    for row in sales_sheet.iter_rows(min_row=2, values_only=True):
        name = str(row[2])
        qty = safe_int(row[3])
        value = safe_float(row[5])  # total sale value
        date_str = str(row[1])
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        except:
            continue
        if from_date and date_obj < from_date:
            continue
        if to_date and date_obj > to_date:
            continue
        if item_filter and item_filter not in name:
            continue
        report_table.insert("", "end", values=("بيع", name, qty, value, date_str))

tk.Button(report_controls, text="تحديث التقرير", command=refresh_report_table).grid(row=2, column=1, pady=6)
tk.Button(report_controls, text="تحميل التقرير (CSV)", command=export_report_csv).grid(row=2, column=2, pady=6)
tk.Button(report_controls, text="تحميل التقرير (Excel)", command=export_report_excel).grid(row=2, column=3, pady=6)

# ------------------- initial refresh -------------------
refresh_inventory_tree()
refresh_purchases_table()
refresh_sales_table()
refresh_combobox_values(sales_combo)

root.mainloop()
